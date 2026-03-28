import os
import requests
import hashlib
import math
import pandas as pd
from datetime import datetime
import customtkinter as ctk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import threading
from pathlib import Path

# ----------------------
# Meses en español
# ----------------------
MESES = [
    "",  # para que enero sea 1
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]

VT_URL = "http://192.168.10.57/vtigercrm/webservice.php"

# Obtener la ruta real del Escritorio del usuario actual (Windows robusto, fallback a ~/Desktop)
def get_desktop_path():
    # Windows: intenta obtener la ruta del Escritorio con la API de Shell
    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes, create_unicode_buffer
            CSIDL_DESKTOPDIRECTORY = 0x0010
            buf = create_unicode_buffer(wintypes.MAX_PATH)
            ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOPDIRECTORY, None, 0, buf)
            p = Path(buf.value)
            if p.exists():
                return p
        except Exception:
            pass
    # Fallback (Windows, macOS, Linux)
    return Path.home() / "Desktop"

# Ruta de salida dinámica basada en el Escritorio del usuario actual
OUTPUT_DIR = str(get_desktop_path() / "SISTEMAS" / "CUMPLEAÑOS")

# ----------------------
# Funciones VTiger
# ----------------------
def vtiger_getchallenge(username):
    r = requests.get(f"{VT_URL}?operation=getchallenge&username={username}")
    j = r.json()
    if not j['success']:
        raise RuntimeError("getchallenge error: " + str(j))
    return j['result']['token']

def vtiger_login(username, accessKey):
    token = vtiger_getchallenge(username)
    key = hashlib.md5((token + accessKey).encode()).hexdigest()
    data = {'operation': 'login', 'username': username, 'accessKey': key}
    r = requests.post(VT_URL, data=data)
    j = r.json()
    if not j['success']:
        raise RuntimeError("login error: " + str(j))
    return j['result']['sessionName']

def vtiger_query(session, query):
    r = requests.get(f"{VT_URL}?operation=query&sessionName={session}&query={query}")
    j = r.json()
    if not j['success']:
        raise RuntimeError("query error: " + str(j))
    return j['result']

def vtiger_count_contacts(session):
    query = "SELECT count(*) FROM Contacts WHERE cf_2040 IN ('VALIDADO','VALIDADO COMPACTADO','PENDIENTE ACTIVACON','SEGUIMIENTO PAGO');"
    result = vtiger_query(session, query)
    return int(result[0]['count']) if result else 0

# ----------------------
# Obtener datos VTiger en memoria
# ----------------------
def obtener_contactos(username, key, progress_callback=None):
    session = vtiger_login(username, key)

    campos = [
        'contact_no', 'firstname', 'lastname',  # Titular
        'cf_1086',                              # Fecha nacimiento Titular
        'cf_910', 'cf_912', 'cf_914',           # Cónyuge
        'cf_926', 'cf_928', 'cf_930',           # Dep 1
        'cf_1022','cf_1024','cf_1026',          # Dep 2
        'cf_1034','cf_1036','cf_1038',          # Dep 3
        'cf_1052','cf_1054','cf_1056',          # Dep 4
        'cf_1100','cf_1102','cf_1104',          # Dep 5
        'cf_2040'                               # Estado KM 2026
    ]

    total = vtiger_count_contacts(session)
    batch = 100
    pages = math.ceil(total / batch)

    datos = []

    for p in range(pages):
        offset = p * batch
        query = f"SELECT {','.join(campos)} FROM Contacts WHERE cf_2040 IN ('VALIDADO','VALIDADO COMPACTADO') LIMIT {offset},{batch};"
        rows = vtiger_query(session, query)

        for r in rows:
            datos.append({
                "ID Contacto": r.get("contact_no"),
                "Nombre Titular": r.get("firstname"),
                "Apellido Titular": r.get("lastname"),
                "Fecha Nacimiento Titular": r.get("cf_1086"),

                "Nombre Cónyuge": r.get("cf_910"),
                "Apellido Cónyuge": r.get("cf_912"),
                "Fecha Nacimiento Cónyuge": r.get("cf_914"),

                "Nombre Dependiente 1": r.get("cf_926"),
                "Apellido Dependiente 1": r.get("cf_928"),
                "Fecha Nacimiento Dependiente 1": r.get("cf_930"),

                "Nombre Dependiente 2": r.get("cf_1022"),
                "Apellido Dependiente 2": r.get("cf_1024"),
                "Fecha Nacimiento Dependiente 2": r.get("cf_1026"),

                "Nombre Dependiente 3": r.get("cf_1034"),
                "Apellido Dependiente 3": r.get("cf_1036"),
                "Fecha Nacimiento Dependiente 3": r.get("cf_1038"),

                "Nombre Dependiente 4": r.get("cf_1052"),
                "Apellido Dependiente 4": r.get("cf_1054"),
                "Fecha Nacimiento Dependiente 4": r.get("cf_1056"),

                "Nombre Dependiente 5": r.get("cf_1100"),
                "Apellido Dependiente 5": r.get("cf_1102"),
                "Fecha Nacimiento Dependiente 5": r.get("cf_1104"),

                "Estado KM 2026": r.get("cf_2040"),
            })

        if progress_callback:
            progress_callback((p + 1) / pages * 100)

    return pd.DataFrame(datos)

# ----------------------
# Filtrar cumpleañeros
# ----------------------
def filtrar_cumpleaneros(username, key, mes, progress_callback=None):
    df = obtener_contactos(username, key, progress_callback)

    cumpleaneros = []

    for _, row in df.iterrows():
        personas = [
            ("Titular", row.get("Nombre Titular"), row.get("Apellido Titular"), row.get("Fecha Nacimiento Titular")),
            ("Cónyuge", row.get("Nombre Cónyuge"), row.get("Apellido Cónyuge"), row.get("Fecha Nacimiento Cónyuge"))
        ]
        for i in range(1, 6):
            personas.append(("Dependiente", row.get(f"Nombre Dependiente {i}"), row.get(f"Apellido Dependiente {i}"), row.get(f"Fecha Nacimiento Dependiente {i}")))

        for tipo, nombre, apellido, fecha in personas:
            if pd.notna(fecha):
                try:
                    fecha_dt = pd.to_datetime(fecha)
                    if fecha_dt.month == mes:
                        cumpleaneros.append([
                            row["ID Contacto"],
                            nombre,
                            apellido,
                            fecha_dt.strftime("%d-%m-%Y"),
                            tipo,
                            fecha_dt.day
                        ])
                except:
                    pass

    df_final = pd.DataFrame(cumpleaneros, columns=[
        "ID Contacto", "Nombre", "Apellido", "Fecha de Nacimiento", "Clasificación", "Día"
    ])
    df_final = df_final.sort_values(by="Día").drop(columns=["Día"])

    mes_nombre = MESES[mes]
    # Asegurarse de que la carpeta de salida exista y construir la ruta completa al archivo
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_file = os.path.join(OUTPUT_DIR, f"CUMPLEAÑOS {mes_nombre}.xlsx")
    df_final.to_excel(output_file, index=False)

    # Formato bonito en Excel
    wb = load_workbook(output_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(output_file)

    if progress_callback:
        progress_callback(100)

    return mes_nombre, len(df_final)

# ----------------------
# Interfaz gráfica
# ----------------------
def ejecutar_app():
    ctk.set_appearance_mode("system")  
    ctk.set_default_color_theme("blue")  

    root = ctk.CTk()
    root.title("Generador de Cumpleaños")
    root.geometry("480x420")

    # Usuario
    label_user = ctk.CTkLabel(root, text="Usuario", font=("Arial", 14))
    label_user.pack(pady=5)
    entry_user = ctk.CTkEntry(root, width=250)
    entry_user.pack(pady=5)

    # Key con botón 👁
    label_key = ctk.CTkLabel(root, text="Access Key", font=("Arial", 14))
    label_key.pack(pady=5)
    frame_key = ctk.CTkFrame(root)
    frame_key.pack(pady=5)
    entry_key = ctk.CTkEntry(frame_key, width=200, show="*")
    entry_key.pack(side="left", padx=5)

    def toggle_key():
        if entry_key.cget("show") == "*":
            entry_key.configure(show="")
            btn_eye.configure(text="🙈")
        else:
            entry_key.configure(show="*")
            btn_eye.configure(text="👁")
    btn_eye = ctk.CTkButton(frame_key, text="👁", width=40, command=toggle_key)
    btn_eye.pack(side="left")

    # Selector de mes
    label_mes = ctk.CTkLabel(root, text="Seleccione el mes", font=("Arial", 14))
    label_mes.pack(pady=10)
    combo = ctk.CTkComboBox(root, values=MESES[1:])
    combo.pack(pady=5)

    # Barra de progreso
    progress = ctk.CTkProgressBar(root, width=300)
    progress.set(0)
    progress.pack(pady=15)

    # Botón
    def generar():
        usuario = entry_user.get().strip()
        clave = entry_key.get().strip()
        mes_nombre = combo.get()

        if not usuario or not clave or not mes_nombre:
            messagebox.showerror("Error", "Debe ingresar usuario, key y mes.")
            return

        mes_num = MESES.index(mes_nombre)

        def run():
            try:
                def update_progress(val):
                    progress.set(val / 100)

                mes_nombre, total = filtrar_cumpleaneros(usuario, clave, mes_num, update_progress)
                messagebox.showinfo("Éxito", f"Archivo generado.\n\nMes: {mes_nombre}\nCumpleañeros: {total}\n\nGuardado en:\n{os.path.join(OUTPUT_DIR, f'CUMPLEAÑOS {mes_nombre}.xlsx')}")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                progress.set(0)

        threading.Thread(target=run).start()

    boton = ctk.CTkButton(root, text="Generar Excel", command=generar)
    boton.pack(pady=20)

    root.mainloop()

# ------------------------
# Ejecutar App
# ------------------------
if __name__ == "__main__":
    ejecutar_app()