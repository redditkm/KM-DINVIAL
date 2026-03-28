import customtkinter as ctk
from tkinter import messagebox
import tkinter as tk
import requests
import re
import os
import glob
from openpyxl import load_workbook
import sys

UPLOAD_URL = "https://links.dinivalsas.com/upload.php"
BASE_URL = "https://links.dinivalsas.com/"

# ======================
# Rutas dinámicas (funciona en .exe y en .py)
# ======================

USER_HOME = os.path.expanduser("~")

ARCHIVO_CUMPLEANOS = os.path.join(
    USER_HOME,
    "Documents",
    "CUMPLEAÑOS.PNG"
)

if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ======================
# Mensajes Titulares
# ======================

MENSAJE_TITULAR_ACTUAL = """En este día especial, queremos agradecerte por confiar en KM DINIVAL INSURANCE como tu asesor en seguros. ¡Feliz cumpleaños!🥳🎊💃🏼  
Para celebrar contigo, hemos cargado 10 puntos a tu tarjeta de regalo, que esperamos que puedas disfrutar en lo que más te guste.🎉🎁  
Recuerda que 100 puntos equivalen a 100 dólares y al completar esta cantidad enviaremos la tarjeta a tu domicilio para que la puedas usar en lo que más te guste. 🙏🏼  


Para cerrar esta celebración, en el siguiente enlace encontrarás una imagen muy especial que preparamos con mucho cariño para ti:
👉 {enlace}


REQ// SE ENVIA MENSAJE DE CUMPLEAÑOS // SE ACTUALIZA EDAD Y PUNTOS
"""

MENSAJE_TITULAR_PASO = """¡Esperamos que hayas tenido un feliz cumpleaños! En este mes tan especial, en nombre de KM DINIVAL INSURANCE, aunque tu cumpleaños ya pasó, queremos festejar contigo y hacerte un regalo especial.  
Hemos cargado 10 puntos a tu tarjeta de regalo, que esperamos que puedas disfrutar en lo que más te guste.🎉🎁  
Recuerda que por cada persona que nos des a conocer y pertenezca a la agencia de la divinidad le contará como 10 puntos más. 🥳  
100 puntos equivalen a 100 dólares y al completar esta cantidad enviaremos la tarjeta a tu domicilio para que la puedas utilizar en lo que más te guste. 🙏🏼  


Para cerrar esta celebración, en el siguiente enlace encontrarás una imagen muy especial que preparamos con mucho cariño para ti:
👉 {enlace}


REQ// SE ENVIA MENSAJE DE CUMPLEAÑOS // SE ACTUALIZA EDAD Y PUNTOS
"""

MENSAJE_FAMILIAR_ACTUAL = """En nombre de toda nuestra agencia KM DINIVAL, queremos desearle un feliz cumpleaños 🎉 lleno de amor, alegría y salud a {nombre_familiar}, 🎊 esperamos disfruten mucho en este día y que todos sus días estén llenos de bendiciones, felicidad y éxito. 🙏🏻


Queremos cerrar esta celebración con un gesto lleno de cariño. En el siguiente enlace encontrarás una imagen muy especial que hemos preparado pensando en esa persona que hace tu vida más hermosa.
👉 {enlace}


REQ // SE ENVIA MENSAJE DE CUMPLEAÑOS A FAMILIAR
"""

MENSAJE_FAMILIAR_PASO = """Aunque el cumpleaños de {nombre_familiar} ya haya pasado, queremos tomar este momento para enviarles un mensaje especial. 🎊 
Les enviamos nuestros mejores deseos y esperamos que el cumpleaños de {nombre_familiar} haya sido memorable y lleno de amor. 🎉 Que el próximo año esté lleno de bendiciones, felicidad y éxito para {nombre_familiar} y para toda su familia. 🥳
Con cariño, KM DINIVAL 🤗


Queremos cerrar esta celebración con un gesto lleno de cariño. En el siguiente enlace encontrarás una imagen muy especial que hemos preparado pensando en esa persona que hace tu vida más hermosa.
👉 {enlace}


REQ // SE ENVIA MENSAJE DE CUMPLEAÑOS A FAMILIAR
"""

# ======================
# Estado Global
# ======================

estado = {
    "dia_actual": None,
    "familiares_lista": [],
    "indice_actual": 0
}

# ======================
# Función para encontrar XLSX
# ======================

def encontrar_xlsx():
    archivos_xlsx = glob.glob(os.path.join(SCRIPT_DIR, "*.xlsx"))
    if archivos_xlsx:
        return archivos_xlsx[0]
    return None

# ======================
# Funciones para leer XLSX
# ======================

def leer_familiares_del_xlsx(dia):
    archivo_xlsx = encontrar_xlsx()
    
    if not archivo_xlsx:
        return None, f"No se encontró ningún archivo .xlsx en:\n{SCRIPT_DIR}\n\nColoca tu archivo Excel en la misma carpeta que este programa."
    
    try:
        wb = load_workbook(archivo_xlsx)
        ws = wb.active
        
        familiares = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue
            
            codigo, nombre, apellido, fecha_nacimiento, tipo_persona = row[0], row[1], row[2], row[3], row[4]
            
            if not tipo_persona or tipo_persona.strip().upper() not in ["CÓNYUGE", "CONYUGE", "DEPENDIENTE"]:
                continue
            
            if fecha_nacimiento:
                try:
                    if isinstance(fecha_nacimiento, str):
                        from datetime import datetime
                        fecha_obj = datetime.strptime(fecha_nacimiento, "%d-%m-%Y")
                    else:
                        fecha_obj = fecha_nacimiento
                    
                    dia_nacimiento = fecha_obj.day
                    
                    if dia_nacimiento == int(dia):
                        nombre_completo = f"{nombre} {apellido}".strip()
                        familiares.append({
                            "nombre": nombre_completo,
                            "tipo": tipo_persona.strip(),
                            "fecha": fecha_obj
                        })
                except:
                    continue
        
        if not familiares:
            return None, f"No se encontraron familiares (cónyuge/dependiente) con cumpleaños el día {dia}"
        
        familiares.reverse()
        return familiares, None
    
    except Exception as e:
        return None, f"Error al leer el XLSX: {e}"

# ======================
# Funciones para subir archivo
# ======================

def subir_archivo(file_path):
    try:
        with open(file_path, "rb") as f:
            files = {"files[]": f}
            data = {"submit": "Subir Archivos"}
            response = requests.post(UPLOAD_URL, files=files, data=data, timeout=5)

        match = re.search(r"href='([^']+)'", response.text)
        if match:
            enlace = match.group(1)
            if not enlace.startswith("http"):
                enlace = BASE_URL + enlace.lstrip("/")
            return enlace
        return None

    except Exception as e:
        messagebox.showerror("Error", f"Error al subir archivo:\n{e}")
        return None

# ======================
# Generar mensajes
# ======================

def generar_mensaje_titular():
    tipo = combo_tipo.get()
    
    if not tipo:
        messagebox.showwarning("Advertencia", "Por favor selecciona Actual o Pasó.")
        return
    
    if not os.path.exists(ARCHIVO_CUMPLEANOS):
        messagebox.showerror("Error", f"No se encontró la imagen en:\n{ARCHIVO_CUMPLEANOS}")
        return

    # SIEMPRE subir imagen (como TITULAR)
    enlace = subir_archivo(ARCHIVO_CUMPLEANOS)
    if not enlace:
        messagebox.showerror("Error", "No se pudo generar el enlace.")
        return

    if tipo == "Actual":
        mensaje = MENSAJE_TITULAR_ACTUAL.format(enlace=enlace)
    else:
        mensaje = MENSAJE_TITULAR_PASO.format(enlace=enlace)

    text_area.delete("1.0", tk.END)
    text_area.insert(tk.END, mensaje)

    btn_titular.configure(fg_color="#28a745")
    root.after(2000, lambda: btn_titular.configure(fg_color="#1a73e8"))

def generar_mensaje_familiar():
    dia = entry_dia.get().strip()
    tipo = combo_tipo.get()
    
    if not dia:
        messagebox.showwarning("Advertencia", "Por favor ingresa un día.")
        return
    
    if not tipo:
        messagebox.showwarning("Advertencia", "Por favor selecciona Actual o Pasó.")
        return
    
    try:
        dia = int(dia)
        if dia < 1 or dia > 31:
            messagebox.showwarning("Advertencia", "El día debe estar entre 1 y 31.")
            return
    except ValueError:
        messagebox.showwarning("Advertencia", "El día debe ser un número.")
        return
    
    # Si es un día diferente, cargar familiares
    if estado["dia_actual"] != dia:
        familiares, error = leer_familiares_del_xlsx(dia)
        
        if error:
            messagebox.showerror("Error", error)
            return
        
        estado["dia_actual"] = dia
        estado["familiares_lista"] = familiares
        estado["indice_actual"] = 0
    
    if estado["indice_actual"] >= len(estado["familiares_lista"]):
        messagebox.showinfo("Información", "Se han generado todos los mensajes para este día.")
        estado["dia_actual"] = None
        estado["indice_actual"] = 0
        label_progreso.configure(text="Familiar 0/0")
        return
    
    if not os.path.exists(ARCHIVO_CUMPLEANOS):
        messagebox.showerror("Error", f"No se encontró la imagen en:\n{ARCHIVO_CUMPLEANOS}")
        return
    
    # SIEMPRE subir imagen (cada vez que se presiona FAMILIAR)
    enlace = subir_archivo(ARCHIVO_CUMPLEANOS)
    if not enlace:
        messagebox.showerror("Error", "No se pudo generar el enlace.")
        return
    
    familiar = estado["familiares_lista"][estado["indice_actual"]]
    nombre_familiar = familiar["nombre"].split()[0]
    
    if tipo == "Actual":
        template = MENSAJE_FAMILIAR_ACTUAL
    else:
        template = MENSAJE_FAMILIAR_PASO
    
    # Usar el enlace nuevo (se sube cada vez)
    mensaje = template.format(
        nombre_familiar=nombre_familiar,
        enlace=enlace
    )
    
    text_area.delete("1.0", tk.END)
    text_area.insert(tk.END, mensaje)
    
    btn_familiar.configure(fg_color="#28a745")
    root.after(2000, lambda: btn_familiar.configure(fg_color="#1a73e8"))
    
    estado["indice_actual"] += 1
    
    total = len(estado["familiares_lista"])
    actual = estado["indice_actual"]
    label_progreso.configure(text=f"Familiar {actual}/{total}")

def reiniciar():
    """Reinicia el estado para comenzar nuevamente"""
    estado["dia_actual"] = None
    estado["familiares_lista"] = []
    estado["indice_actual"] = 0
    
    combo_tipo.set("")
    entry_dia.delete(0, tk.END)
    entry_dia.insert(0, "00")
    text_area.delete("1.0", tk.END)
    label_progreso.configure(text="Familiar 0/0")
    
    messagebox.showinfo("Reiniciado", "El generador ha sido reiniciado. Puedes comenzar nuevamente.")

def cerrar_combo_si_abierto():
    try:
        combo_tipo._dropdown_menu.dismiss()
    except:
        pass

# ======================
# Configurar tema
# ======================

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ======================
# Interfaz
# ======================

root = ctk.CTk()
root.title("🎂 Mensajes de Cumpleaños KM DINIVAL 🎂")
root.geometry("700x650")
root.resizable(True, True)

root.grid_rowconfigure(3, weight=1)
root.grid_columnconfigure(0, weight=1)

# ===== TITULO =====
titulo = ctk.CTkLabel(
    root,
    text="Generador de Mensajes",
    font=("Segoe UI", 22, "bold"),
    text_color="#1a73e8"
)
titulo.grid(row=0, column=0, pady=10, padx=15, sticky="ew")

# ===== FRAME CONFIG =====
frame_config = ctk.CTkFrame(root, fg_color="transparent")
frame_config.grid(row=1, column=0, pady=8, padx=15, sticky="ew")

frame_config.grid_columnconfigure(0, weight=1)
frame_config.grid_columnconfigure(1, weight=1)

label_tipo = ctk.CTkLabel(
    frame_config,
    text="Tipo de cumpleaños:",
    font=("Segoe UI", 11, "bold"),
    text_color="#333"
)
label_tipo.grid(row=0, column=0, sticky="w", padx=0, pady=(0, 5))

combo_tipo = ctk.CTkComboBox(
    frame_config,
    values=["Actual", "Pasó"],
    font=("Segoe UI", 11),
    height=32
)
combo_tipo.grid(row=1, column=0, sticky="ew", padx=(0, 5))
combo_tipo.set("")

label_dia = ctk.CTkLabel(
    frame_config,
    text="Día:",
    font=("Segoe UI", 11, "bold"),
    text_color="#333"
)
label_dia.grid(row=0, column=1, sticky="w", padx=0, pady=(0, 5))

entry_dia = ctk.CTkEntry(
    frame_config,
    font=("Segoe UI", 11),
    height=32
)
entry_dia.grid(row=1, column=1, sticky="ew", padx=(5, 0))
entry_dia.insert(0, "00")

combo_tipo.bind("<FocusOut>", lambda e: cerrar_combo_si_abierto())
entry_dia.bind("<FocusIn>", lambda e: cerrar_combo_si_abierto())

# ===== FRAME BOTONES =====
frame_botones = ctk.CTkFrame(root, fg_color="transparent")
frame_botones.grid(row=2, column=0, pady=5, padx=15, sticky="ew")

frame_botones.grid_columnconfigure(0, weight=1)
frame_botones.grid_columnconfigure(1, weight=1)

btn_titular = ctk.CTkButton(
    frame_botones,
    text="TITULAR",
    command=generar_mensaje_titular,
    font=("Segoe UI", 11, "bold"),
    height=36,
    fg_color="#1a73e8",
    hover_color="#0d5dbf",
    text_color="white",
    corner_radius=6
)
btn_titular.grid(row=0, column=0, sticky="ew", padx=(0, 5))

btn_familiar = ctk.CTkButton(
    frame_botones,
    text="FAMILIAR",
    command=generar_mensaje_familiar,
    font=("Segoe UI", 11, "bold"),
    height=36,
    fg_color="#1a73e8",
    hover_color="#0d5dbf",
    text_color="white",
    corner_radius=6
)
btn_familiar.grid(row=0, column=1, sticky="ew", padx=(5, 0))

label_progreso = ctk.CTkLabel(
    frame_botones,
    text="Familiar 0/0",
    font=("Segoe UI", 9),
    text_color="#666"
)
label_progreso.grid(row=1, column=1, sticky="e", pady=(4, 0))

# ===== FRAME MENSAJE =====
frame_mensaje = ctk.CTkFrame(root, fg_color="transparent")
frame_mensaje.grid(row=3, column=0, pady=8, padx=15, sticky="nsew")

frame_mensaje.grid_rowconfigure(1, weight=1)
frame_mensaje.grid_columnconfigure(0, weight=1)

label_mensaje = ctk.CTkLabel(
    frame_mensaje,
    text="Mensaje generado:",
    font=("Segoe UI", 11, "bold"),
    text_color="#333"
)
label_mensaje.grid(row=0, column=0, sticky="w", pady=(0, 5))

text_area = tk.Text(
    frame_mensaje,
    font=("Segoe UI", 12),
    wrap=tk.WORD,
    bg="#ffffff",
    fg="#222222",
    relief=tk.FLAT,
    bd=1,
    borderwidth=1,
    highlightthickness=1,
    highlightcolor="#d0d0d0",
    padx=12,
    pady=12
)
text_area.grid(row=1, column=0, sticky="nsew")

scrollbar = tk.Scrollbar(frame_mensaje, command=text_area.yview)
scrollbar.grid(row=1, column=1, sticky="ns")
text_area.config(yscrollcommand=scrollbar.set)

# ===== FRAME BOTÓN REINICIAR =====
frame_reiniciar = ctk.CTkFrame(root, fg_color="transparent")
frame_reiniciar.grid(row=4, column=0, pady=8, padx=15, sticky="ew")

btn_reiniciar = ctk.CTkButton(
    frame_reiniciar,
    text="REINICIAR",
    command=reiniciar,
    font=("Segoe UI", 11, "bold"),
    height=36,
    fg_color="#ff6b6b",
    hover_color="#e63946",
    text_color="white",
    corner_radius=6
)
btn_reiniciar.pack(fill="x")

root.mainloop()